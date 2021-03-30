from openpyxl import Workbook, load_workbook
import random

def create_table(excel_filename):
    wb = Workbook()
    ws = wb["Sheet"]
    ws.title = "Random table 12x12"
    for row in ws.iter_rows(min_row=1,
                            min_col=1,
                            max_col=12,
                            max_row=12):
        for cell in row:
            cell.value = random.randint(1, 100)
    wb.save(excel_filename)
    return 1

excel_filename = 'random.xlsx'
create_table(excel_filename)